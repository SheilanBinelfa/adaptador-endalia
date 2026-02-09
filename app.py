import streamlit as st
import openpyxl
from io import BytesIO
from zipfile import ZipFile, ZIP_DEFLATED
import xml.etree.ElementTree as ET
import re
from datetime import datetime, time, date
from copy import copy
import unicodedata

SPREADSHEET_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
ET.register_namespace('', SPREADSHEET_NS)
ET.register_namespace('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships')
ET.register_namespace('mc', 'http://schemas.openxmlformats.org/markup-compatibility/2006')
ET.register_namespace('x14ac', 'http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac')
ET.register_namespace('xr', 'http://schemas.microsoft.com/office/spreadsheetml/2014/revision')
ET.register_namespace('xr2', 'http://schemas.microsoft.com/office/spreadsheetml/2015/revision2')
ET.register_namespace('xr3', 'http://schemas.microsoft.com/office/spreadsheetml/2016/revision3')
ET.register_namespace('xr6', 'http://schemas.microsoft.com/office/spreadsheetml/2014/revision6')
ET.register_namespace('xr10', 'http://schemas.microsoft.com/office/spreadsheetml/2018/revision10')


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# XML / ZIP â€” Preservar validaciones
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def extract_all_validations(file_bytes):
    validations = {}
    ns = '{' + SPREADSHEET_NS + '}'
    with ZipFile(BytesIO(file_bytes), 'r') as zf:
        for name in zf.namelist():
            if name.startswith('xl/worksheets/') and name.endswith('.xml'):
                tree = ET.fromstring(zf.read(name))
                dv_node = tree.find(ns + 'dataValidations')
                if dv_node is not None:
                    validations[name] = ET.tostring(dv_node, encoding='unicode')
    return validations


def expand_sqref(sqref, max_row):
    parts = sqref.split(' ')
    new_parts = []
    for part in parts:
        if ':' in part:
            start, end = part.split(':')
            col_end = re.match(r'([A-Z]+)', end)
            row_end = re.search(r'(\d+)', end)
            if col_end and row_end:
                old_row = int(row_end.group(1))
                if max_row > old_row:
                    new_parts.append(f"{start}:{col_end.group(1)}{max_row}")
                else:
                    new_parts.append(part)
            else:
                new_parts.append(part)
        else:
            new_parts.append(part)
    return ' '.join(new_parts)


def patch_zip_with_validations(output_bytes, original_bytes):
    ns = '{' + SPREADSHEET_NS + '}'
    original_zip = ZipFile(BytesIO(original_bytes), 'r')
    output_zip_in = ZipFile(BytesIO(output_bytes), 'r')
    original_sheets = {}
    for name in original_zip.namelist():
        if name.startswith('xl/worksheets/') and name.endswith('.xml'):
            original_sheets[name] = original_zip.read(name)
    result_buffer = BytesIO()
    result_zip = ZipFile(result_buffer, 'w', ZIP_DEFLATED)
    processed = set()
    for item in output_zip_in.namelist():
        data = output_zip_in.read(item)
        processed.add(item)
        if item in original_sheets:
            try:
                output_tree = ET.fromstring(data)
                original_tree = ET.fromstring(original_sheets[item])
                for dv in output_tree.findall(ns + 'dataValidations'):
                    output_tree.remove(dv)
                original_dv = original_tree.find(ns + 'dataValidations')
                if original_dv is not None:
                    output_sd = output_tree.find(ns + 'sheetData')
                    if output_sd is not None:
                        all_rows = output_sd.findall(ns + 'row')
                        if all_rows:
                            max_row = max(int(r.get('r', '1')) for r in all_rows)
                            for dv_item in original_dv.findall(ns + 'dataValidation'):
                                sqref = dv_item.get('sqref', '')
                                dv_item.set('sqref', expand_sqref(sqref, max_row))
                    insert_before = [
                        ns + 'pageMargins', ns + 'pageSetup', ns + 'headerFooter',
                        ns + 'drawing', ns + 'legacyDrawing', ns + 'tableParts', ns + 'extLst',
                    ]
                    inserted = False
                    for tag in insert_before:
                        target = output_tree.find(tag)
                        if target is not None:
                            idx = list(output_tree).index(target)
                            output_tree.insert(idx, original_dv)
                            inserted = True
                            break
                    if not inserted:
                        output_tree.append(original_dv)
                original_ext = original_tree.find(ns + 'extLst')
                if original_ext is not None:
                    for ext in output_tree.findall(ns + 'extLst'):
                        output_tree.remove(ext)
                    output_tree.append(original_ext)
                data = b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                data += ET.tostring(output_tree, encoding='unicode').encode('utf-8')
            except ET.ParseError:
                pass
        result_zip.writestr(item, data)
    for item in original_zip.namelist():
        if item not in processed:
            result_zip.writestr(item, original_zip.read(item))
    original_zip.close()
    output_zip_in.close()
    result_zip.close()
    return result_buffer.getvalue()


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# NormalizaciÃ³n de nombres
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def remove_accents(text):
    """Quita acentos: Ã©->e, Ã±->n, etc."""
    nfkd = unicodedata.normalize('NFKD', text)
    return ''.join(c for c in nfkd if not unicodedata.category(c).startswith('M'))


def normalize_name(name):
    """Normaliza nombre para matching robusto."""
    if name is None:
        return ''
    s = str(name).strip()
    s = s.lower()
    s = remove_accents(s)
    s = re.sub(r'\s+', ' ', s)
    s = s.strip()
    return s


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# Lectura de Registros de Tramos
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def read_tramos(wb):
    ws = wb.active
    col_map = {}
    keywords = {
        'fecha': ['fecha'],
        'hora_inicio': ['hora inicio'],
        'hora_fin': ['hora fin'],
        'duracion': ['duracion', 'duraciÃ³n'],
        'tipo_tramo': ['tipo de tramo', 'tipo de tra', 'tipo tramo'],
        'empleado': ['empleado'],
        'validado': ['validado'],
        'timezone_name': ['timezonename'],
        'timezone_offset': ['timezoneoffset'],
    }
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val is None:
            continue
        val_lower = str(val).strip().lower()
        for key, terms in keywords.items():
            for term in terms:
                if term in val_lower and key not in col_map:
                    col_map[key] = col
                    break

    if 'empleado' not in col_map:
        return None, col_map, "No se encontro columna 'Empleado' en los registros."

    tramos = []
    for row in range(2, ws.max_row + 1):
        emp = ws.cell(row=row, column=col_map['empleado']).value
        if emp is None or str(emp).strip() == '':
            continue
        tramo = {'row': row, 'empleado': str(emp).strip()}
        for key in ['fecha', 'hora_inicio', 'hora_fin', 'tipo_tramo', 'timezone_name']:
            if key in col_map:
                tramo[key] = ws.cell(row=row, column=col_map[key]).value
        tramos.append(tramo)

    return tramos, col_map, None


def is_missing_hora_fin(hora_fin):
    if hora_fin is None:
        return True
    if isinstance(hora_fin, time):
        return hora_fin == time(0, 0)
    if isinstance(hora_fin, datetime):
        return hora_fin.hour == 0 and hora_fin.minute == 0
    s = str(hora_fin).strip()
    return s == '' or s in ('00:00', '0:00', '00:00:00')


def fmt_time(val):
    if val is None:
        return ''
    if isinstance(val, time):
        return val.strftime('%H:%M')
    if isinstance(val, datetime):
        return val.strftime('%H:%M')
    return str(val)


def fmt_date(val):
    if val is None:
        return ''
    if isinstance(val, (datetime, date)):
        return val.strftime('%d/%m/%Y')
    return str(val)


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# DetecciÃ³n de columnas de la plantilla
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def find_plantilla_columns(ws, header_row=1):
    cols = {}
    # Recorrer todas las columnas y asignar por header exacto/parcial
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is None:
            continue
        h = str(val).strip().lower()

        if 'doc. identificador' in h or h == 'nif' or h == 'dni':
            cols.setdefault('nif', col)
        elif 'codigo empleado' in h or 'cÃ³digo empleado' in h:
            cols.setdefault('codigo', col)
        elif h == 'empleado':
            # Solo match EXACTO para evitar confundir con "CÃ³digo empleado"
            cols.setdefault('empleado', col)
        elif 'fecha de referencia' in h:
            cols.setdefault('fecha_ref', col)
        elif 'zona horaria' in h:
            cols.setdefault('zona', col)
        elif h == 'inicio':
            cols.setdefault('inicio', col)
        elif h == 'fin':
            cols.setdefault('fin', col)
        elif 'tipo de tramo' in h:
            cols.setdefault('tipo_tramo', col)
        elif 'sobrescritura' in h:
            cols.setdefault('sobrescritura', col)

    return cols


def copy_cell_style(src, tgt):
    if src.has_style:
        tgt.font = copy(src.font)
        tgt.border = copy(src.border)
        tgt.fill = copy(src.fill)
        tgt.number_format = src.number_format
        tgt.protection = copy(src.protection)
        tgt.alignment = copy(src.alignment)


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# Motor de conciliaciÃ³n
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def conciliar(wb_template, tramos, template_sheet_name, header_row, zona_default, sobrescritura_default):
    if template_sheet_name and template_sheet_name in wb_template.sheetnames:
        ws = wb_template[template_sheet_name]
    else:
        ws = wb_template.active

    cols = find_plantilla_columns(ws, header_row)
    debug_info = {'cols': cols}

    if 'empleado' not in cols:
        return 0, 0, [], debug_info, "No se encontro columna 'Empleado' en la plantilla."

    # 1. Leer empleados de la plantilla
    plantilla_emps = []
    for row in range(header_row + 1, ws.max_row + 1):
        emp_val = ws.cell(row=row, column=cols['empleado']).value
        if emp_val is None or str(emp_val).strip() == '':
            continue
        plantilla_emps.append({
            'row': row,
            'nombre': str(emp_val).strip(),
            'nombre_norm': normalize_name(emp_val),
            'nif': ws.cell(row=row, column=cols.get('nif', 1)).value,
            'codigo': ws.cell(row=row, column=cols.get('codigo', 2)).value,
        })

    debug_info['plantilla_count'] = len(plantilla_emps)
    debug_info['plantilla_names_sample'] = [p['nombre_norm'] for p in plantilla_emps[:5]]

    # 2. Agrupar tramos por empleado
    tramos_by_emp = {}
    for t in tramos:
        key = normalize_name(t['empleado'])
        if key not in tramos_by_emp:
            tramos_by_emp[key] = []
        tramos_by_emp[key].append(t)

    debug_info['tramos_keys_sample'] = list(tramos_by_emp.keys())[:5]

    # 3. Match y construir filas de salida
    output_rows = []
    matched_emp_norms = set()
    unmatched = []

    for tramo_emp_norm, emp_tramos in tramos_by_emp.items():
        # Buscar match exacto
        found = None
        for pe in plantilla_emps:
            if pe['nombre_norm'] == tramo_emp_norm:
                found = pe
                break

        # Match parcial si no hay exacto
        if found is None:
            for pe in plantilla_emps:
                if tramo_emp_norm in pe['nombre_norm'] or pe['nombre_norm'] in tramo_emp_norm:
                    found = pe
                    break

        if found is None:
            for t in emp_tramos:
                unmatched.append(t['empleado'])
            continue

        matched_emp_norms.add(found['nombre_norm'])

        for t in emp_tramos:
            output_rows.append({
                'nif': found['nif'],
                'codigo': found['codigo'],
                'nombre': found['nombre'],
                'fecha_ref': t.get('fecha'),
                'zona': t.get('timezone_name') or zona_default,
                'inicio': t.get('hora_inicio'),
                'fin': t.get('hora_fin'),
                'tipo_tramo': t.get('tipo_tramo'),
                'sobrescritura': sobrescritura_default,
            })

    debug_info['matched_employees'] = len(matched_emp_norms)
    debug_info['output_rows'] = len(output_rows)

    # 4. Contar eliminados (empleados de plantilla sin tramos)
    removed = 0
    for pe in plantilla_emps:
        if pe['nombre_norm'] not in matched_emp_norms:
            removed += 1

    # 5. Guardar estilos de la primera fila de datos
    style_row = header_row + 1
    styles = {}
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        styles[c] = ws.cell(row=style_row, column=c)

    # 6. Limpiar filas de datos
    old_max_row = ws.max_row
    for row in range(header_row + 1, old_max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).value = None

    # 7. Escribir filas de salida
    for i, item in enumerate(output_rows):
        row = header_row + 1 + i

        # Copiar estilo
        for c in range(1, max_col + 1):
            copy_cell_style(styles[c], ws.cell(row=row, column=c))

        # Escribir datos
        if 'nif' in cols:
            ws.cell(row=row, column=cols['nif']).value = item['nif']
        if 'codigo' in cols:
            ws.cell(row=row, column=cols['codigo']).value = item['codigo']
        if 'empleado' in cols:
            ws.cell(row=row, column=cols['empleado']).value = item['nombre']
        if 'fecha_ref' in cols:
            ws.cell(row=row, column=cols['fecha_ref']).value = item['fecha_ref']
        if 'zona' in cols:
            ws.cell(row=row, column=cols['zona']).value = item['zona']
        if 'inicio' in cols:
            ws.cell(row=row, column=cols['inicio']).value = item['inicio']
        if 'fin' in cols:
            ws.cell(row=row, column=cols['fin']).value = item['fin']
        if 'tipo_tramo' in cols:
            ws.cell(row=row, column=cols['tipo_tramo']).value = item['tipo_tramo']
        if 'sobrescritura' in cols:
            ws.cell(row=row, column=cols['sobrescritura']).value = item['sobrescritura']

    return len(output_rows), removed, list(set(unmatched)), debug_info, None


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# STREAMLIT UI
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def main():
    st.set_page_config(page_title="Conciliador Endalia", page_icon="ğŸ“Š", layout="wide")
    st.title("ğŸ“Š Conciliador de Tramos â†’ Plantilla Endalia")
    st.markdown("Importa tramos de fichaje a la plantilla Endalia preservando desplegables y formato.")
    st.divider()

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("1ï¸âƒ£ Registros de Tramos")
        st.caption("Excel con fichajes (Fecha, Hora inicio, Hora fin, Empleado...)")
        tramos_file = st.file_uploader("Sube registros de tramos", type=['xlsx'], key='tramos')
    with col2:
        st.subheader("2ï¸âƒ£ Plantilla Endalia")
        st.caption("Plantilla .xlsx con desplegables a preservar.")
        plantilla_file = st.file_uploader("Sube la plantilla", type=['xlsx'], key='plantilla')

    st.divider()

    with st.expander("âš™ï¸ Configuracion"):
        header_row = st.number_input("Fila de encabezados en la plantilla", min_value=1, value=1)
        template_sheet = st.text_input("Hoja de la plantilla", value="Registros de jornada")
        zona_default = st.selectbox("Zona horaria por defecto", [
            "(UTC+01:00) Bruselas, Copenhague, Madrid, ParÃ­s",
            "(UTC+00:00) DublÃ­n, Edimburgo, Lisboa, Londres",
            "(UTC+02:00) Atenas, Bucarest, Estambul",
        ])
        sobrescritura_default = st.selectbox("Sobrescritura por defecto", ["SÃ­", "No"])

    if not tramos_file or not plantilla_file:
        st.info("ğŸ‘† Sube ambos archivos para continuar.")
        return

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # PASO 1: Leer tramos
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    tramos_bytes = tramos_file.read()
    wb_tramos = openpyxl.load_workbook(BytesIO(tramos_bytes), data_only=True)
    tramos, tramos_cols, error = read_tramos(wb_tramos)

    if error:
        st.error(f"âŒ {error}")
        return

    st.success(f"âœ… {len(tramos)} tramos leidos del archivo de registros.")

    # Mostrar columnas detectadas en tramos
    with st.expander("ğŸ” Columnas detectadas en registros"):
        ws_t = wb_tramos.active
        for key, idx in tramos_cols.items():
            header = ws_t.cell(row=1, column=idx).value
            st.write(f"**{key}** â†’ Col {idx} = \"{header}\"")

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # PASO 2: Tramos sin Hora Fin
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    tramos_sin_fin = [t for t in tramos if is_missing_hora_fin(t.get('hora_fin'))]
    tramos_con_fin = [t for t in tramos if not is_missing_hora_fin(t.get('hora_fin'))]

    if tramos_sin_fin:
        st.warning(f"âš ï¸ **{len(tramos_sin_fin)}** tramos sin Hora Fin. Introduce la hora fin:")
        st.markdown("---")
        hora_fin_inputs = {}
        for i, t in enumerate(tramos_sin_fin):
            c1, c2, c3, c4 = st.columns([3, 2, 2, 2])
            with c1:
                st.write(f"**{t['empleado']}**")
            with c2:
                st.write(f"ğŸ“… {fmt_date(t.get('fecha'))}")
            with c3:
                st.write(f"ğŸ• Inicio: {fmt_time(t.get('hora_inicio'))}")
            with c4:
                hora_fin_inputs[i] = st.time_input("Hora Fin", value=time(17, 0), key=f"hf_{i}")
        st.markdown("---")

        for i, t in enumerate(tramos_sin_fin):
            t['hora_fin'] = hora_fin_inputs[i]

        all_tramos = tramos_con_fin + tramos_sin_fin
    else:
        all_tramos = tramos
        st.info("âœ… Todos los tramos tienen Hora Fin.")

    # Aplicar defaults
    for t in all_tramos:
        if not t.get('timezone_name'):
            t['timezone_name'] = zona_default

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # PASO 3: Preview
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    with st.expander(f"ğŸ‘€ Vista previa: {len(all_tramos)} tramos"):
        for t in all_tramos[:30]:
            st.write(
                f"**{t['empleado']}** | "
                f"{fmt_date(t.get('fecha'))} | "
                f"{fmt_time(t.get('hora_inicio'))} â†’ {fmt_time(t.get('hora_fin'))} | "
                f"{t.get('tipo_tramo', '-')}"
            )
        if len(all_tramos) > 30:
            st.caption(f"... y {len(all_tramos) - 30} mas")

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # PASO 4: Procesar
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    if st.button("ğŸš€ Generar Plantilla", type="primary", use_container_width=True):
        with st.spinner("Procesando..."):
            try:
                plantilla_bytes = plantilla_file.read()

                # Validaciones originales
                original_validations = extract_all_validations(plantilla_bytes)
                st.info(f"ğŸ” {len(original_validations)} hoja(s) con validaciones detectadas.")

                # Cargar plantilla
                wb_template = openpyxl.load_workbook(BytesIO(plantilla_bytes), data_only=False)

                # Conciliar
                written, removed, unmatched, debug, err = conciliar(
                    wb_template, all_tramos, template_sheet, header_row,
                    zona_default, sobrescritura_default
                )

                if err:
                    st.error(f"âŒ {err}")
                    return

                # Debug info
                with st.expander("ğŸ› Info de depuracion"):
                    st.json({
                        'columnas_plantilla': {k: v for k, v in debug['cols'].items()},
                        'empleados_en_plantilla': debug['plantilla_count'],
                        'nombres_plantilla_sample': debug.get('plantilla_names_sample', []),
                        'nombres_tramos_sample': debug.get('tramos_keys_sample', []),
                        'empleados_matched': debug['matched_employees'],
                        'filas_output': debug['output_rows'],
                    })

                st.success(f"âœ… **{written}** filas escritas en la plantilla.")
                if removed:
                    st.info(f"ğŸ—‘ï¸ **{removed}** empleados sin tramos eliminados.")
                if unmatched:
                    with st.expander(f"âš ï¸ {len(unmatched)} empleados sin match"):
                        for name in sorted(unmatched):
                            st.write(f"- {name}")

                # Guardar
                output_buffer = BytesIO()
                wb_template.save(output_buffer)
                output_bytes = output_buffer.getvalue()

                # Parche XML
                st.info("ğŸ”§ Re-inyectando validaciones...")
                output_bytes = patch_zip_with_validations(output_bytes, plantilla_bytes)

                # Verificacion
                final_vals = extract_all_validations(output_bytes)
                if final_vals:
                    st.success(f"ğŸ¯ Verificacion OK: {len(final_vals)} hoja(s) con desplegables.")
                else:
                    st.warning("âš ï¸ No se detectaron validaciones en el archivo final.")

                # Descarga
                st.divider()
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                st.download_button(
                    label="ğŸ“¥ Descargar Plantilla Completada",
                    data=output_bytes,
                    file_name=f"endalia_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet",
                    type="primary",
                    use_container_width=True,
                )

            except Exception as e:
                st.error(f"âŒ Error: {str(e)}")
                st.exception(e)


main()
